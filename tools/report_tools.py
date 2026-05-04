import anthropic
import os
import json
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import datetime


class ReportTools:
    def __init__(self):
        self.client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        self.model  = 'claude-haiku-4-5-20251001'

    # ------------------------------------------------------------------ #
    # HELPERS                                                              #
    # ------------------------------------------------------------------ #

    def _clean_sql(self, sql):
        """Remove markdown link formatting from SQL"""
        # handles [c.id](http://c.id) → c.id
        # handles [cg.car](http://cg.car)_id → cg.car_id
        # step 1: replace [text](url) keeping trailing non-space chars
        cleaned = re.sub(r'\[([^\]]+)\]\(http[^\)]*\)', r'\1', sql)
        return cleaned

    # ------------------------------------------------------------------ #
    # AI — UNDERSTAND REPORT REQUEST                                       #
    # ------------------------------------------------------------------ #

    def understand_report_request(self, user_request):
        schema = """
Base de données MySQL: airliquide_flotte

TABLE cars: id, plate_number, brand, current_cg_id, status(active/inactive/maintenance/retired), acquisition_date
TABLE carte_grises: id, car_id, model, year, owner_name, chassis_number, puissance_fiscale, carburant(essence/diesel/hybride/electrique), registration_date, expiration_date
TABLE employees: id, nom, prenom, email, telephone, poste, departement, status(active/inactive)
TABLE car_assignments: id, car_id, employee_id, start_date, end_date(NULL=actif), notes
TABLE maintenance_records: id, car_id, part_id, garage_id, done_at, km_at_service, next_due_date, next_due_km, status(pending_facture/done/cancelled)
TABLE maintenance_alerts: id, car_id, part_id, alert_type(km/date/both), due_date, due_km, status(open/acknowledged/closed)
TABLE car_parts: id, name, category, alert_km_interval, alert_month_interval
TABLE car_km: id, car_id, km, recorded_at
TABLE factures: id, type(maintenance/sinistre), reference_id, car_id, num_facture, date_facture, montant_ht, montant_ttc, tva
TABLE sinistres: id, car_id, employee_id, date_sinistre, type(accident/vol/vandalisme/autre), description, montant_reparation, status(ouvert/en_cours/cloture)
TABLE insurances: id, car_id, insurer, policy_number, start_date, end_date, status(active/expired/pending/cancelled)
TABLE vignettes: id, car_id, year, expiration_date, montant, status(active/expired/pending)
TABLE visite_technique: id, car_id, expiration_date, montant, status(active/expired/pending)
TABLE garages: id, name, type(dealership/independent), phone, address, status(active/inactive)
TABLE carburant_expenses: id, car_id, employee_id, periode, litres, montant_ttc
"""

        prompt = f"""Tu es un expert SQL MySQL. Génère une requête SQL SELECT pour cette demande.

SCHEMA:
{schema}

DEMANDE: "{user_request}"

REGLES STRICTES:
- Retourne UNIQUEMENT le JSON ci-dessous, sans markdown, sans explication, sans backticks
- SELECT valide MySQL uniquement
- LIMIT 100 maximum
- JOIN avec carte_grises pour avoir model et year du véhicule
- Pour employé actuel d'un véhicule: JOIN car_assignments ca ON ca.car_id=c.id AND ca.end_date IS NULL
- Écrire le SQL en texte brut: écrire c.id et NON [c.id](http://c.id)
- Ne JAMAIS utiliser de syntaxe markdown dans le SQL
- NE PAS inclure les colonnes: id, car_id, employee_id, part_id, garage_id, current_cg_id, reference_id
- Inclure uniquement des colonnes lisibles (noms, plaques, dates, montants)
- Aliaser les colonnes avec des noms français: AS "Immatriculation", AS "Modèle", etc.

FORMAT DE RÉPONSE (JSON pur, pas de markdown):
{{"title": "titre court en français", "sql": "SELECT ..."}}"""

        try:
            response = self.client.messages.create(
                model      = self.model,
                max_tokens = 1000,
                messages   = [{'role': 'user', 'content': prompt}]
            )
            raw = response.content[0].text.strip()
            print(f'Claude raw: {raw}')

            # clean markdown code blocks
            if '```json' in raw:
                raw = raw.split('```json')[1].split('```')[0].strip()
            elif '```' in raw:
                raw = raw.split('```')[1].split('```')[0].strip()

            result = json.loads(raw)

            if 'sql' in result:
                result['sql'] = self._clean_sql(result['sql'])
                print(f'Cleaned SQL: {result["sql"]}')

            return result

        except json.JSONDecodeError as e:
            print(f'JSON parse error: {e} — raw: {raw}')
            return {'error': 'Erreur de parsing de la réponse'}
        except Exception as e:
            print(f'Claude API error: {e}')
            return {'error': f'Erreur API: {str(e)}'}

    # ------------------------------------------------------------------ #
    # PDF GENERATION                                                       #
    # ------------------------------------------------------------------ #

    def generate_pdf(self, title, description, headers, rows, filters_text=''):
        buffer = BytesIO()

        # use A3 for wide reports
        page_size = landscape(A3) if len(headers) > 8 else landscape(A4)
        doc = SimpleDocTemplate(
            buffer, pagesize=page_size,
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm
        )

        styles   = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, textColor=colors.HexColor('#0d6efd'),
            spaceAfter=4, alignment=TA_CENTER
        )
        sub_style = ParagraphStyle(
            'SubTitle', parent=styles['Normal'],
            fontSize=10, textColor=colors.HexColor('#6c757d'),
            alignment=TA_CENTER, spaceAfter=2
        )

        elements.append(Paragraph('AIR LIQUIDE TUNISIA', title_style))
        elements.append(Paragraph('Système de Gestion de Flotte', sub_style))
        elements.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#0d6efd')))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(title, ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'],
            fontSize=14, textColor=colors.HexColor('#2d3748'), alignment=TA_LEFT
        )))
        if description:
            elements.append(Paragraph(description, sub_style))
        elements.append(Paragraph(
            f"Généré le {datetime.date.today().strftime('%d/%m/%Y')}{' | ' + filters_text if filters_text else ''}",
            ParagraphStyle('Meta', parent=styles['Normal'], fontSize=9,
                          textColor=colors.HexColor('#adb5bd'), spaceAfter=10)
        ))
        elements.append(Spacer(1, 0.3*cm))

        if rows:
            page_width = page_size[0] - 3*cm
            col_count  = len(headers)

            # calculate proportional column widths
            col_widths = []
            for i, header in enumerate(headers):
                max_len = len(str(header))
                for row in rows[:20]:
                    if i < len(row):
                        max_len = max(max_len, len(str(row[i])))
                col_widths.append(max_len)

            total_len  = sum(col_widths) or 1
            col_widths = [max((w / total_len) * page_width, 1*cm) for w in col_widths]

            table_data = [headers] + rows
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#0d6efd')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,0), 7),
                ('ALIGN',         (0,0), (-1,0), 'CENTER'),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fc')]),
                ('FONTSIZE',      (0,1), (-1,-1), 7),
                ('ALIGN',         (0,1), (-1,-1), 'CENTER'),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
                ('ROWHEIGHT',     (0,0), (-1,-1), 18),
                ('TOPPADDING',    (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph('Aucune donnée disponible.', styles['Normal']))

        elements.append(Spacer(1, 0.5*cm))
        elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#dee2e6')))
        elements.append(Paragraph(
            f'ALT Fleet Management — {datetime.date.today().year} — {len(rows)} enregistrement(s)',
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8,
                          textColor=colors.HexColor('#adb5bd'), alignment=TA_CENTER)
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------ #
    # EXCEL GENERATION                                                     #
    # ------------------------------------------------------------------ #

    def generate_excel(self, title, headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_fill  = PatternFill('solid', fgColor='0D6EFD')
        title_fill   = PatternFill('solid', fgColor='EEF4FF')
        alt_fill     = PatternFill('solid', fgColor='F8F9FC')
        border_style = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6')
        )

        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'] = 'AIR LIQUIDE TUNISIA — Système de Gestion de Flotte'
        ws['A1'].font      = Font(bold=True, size=14, color='0D6EFD')
        ws['A1'].fill      = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        ws['A2'] = title
        ws['A2'].font      = Font(bold=True, size=12, color='2D3748')
        ws['A2'].fill      = title_fill
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25

        ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
        ws['A3'] = f"Généré le {datetime.date.today().strftime('%d/%m/%Y')} — {len(rows)} enregistrement(s)"
        ws['A3'].font      = Font(size=9, color='6C757D', italic=True)
        ws['A3'].fill      = title_fill
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 8

        for col, header in enumerate(headers, 1):
            cell           = ws.cell(row=5, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFF', size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border_style
        ws.row_dimensions[5].height = 25

        for row_idx, row in enumerate(rows, 6):
            fill = PatternFill('solid', fgColor='FFFFFF') if row_idx % 2 == 0 else alt_fill
            for col, value in enumerate(row, 1):
                cell           = ws.cell(row=row_idx, column=col, value=value)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = border_style
                cell.font      = Font(size=9)
            ws.row_dimensions[row_idx].height = 20

        for col in range(1, len(headers) + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in ws.iter_rows(min_row=5, max_col=col, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 4, 12), 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer